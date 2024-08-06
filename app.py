from flask import Flask, request, jsonify, render_template
# Flask: Framework web para crear aplicaciones web.
# request: Módulo de Flask para manejar solicitudes HTTP.
# jsonify: Módulo de Flask para convertir datos en formato JSON.
# render_template: Módulo de Flask para renderizar plantillas HTML.

import random
# random: Módulo para generar números y secuencias aleatorias.

import nltk
# nltk: Biblioteca para el procesamiento de lenguaje natural.

from nltk.stem import WordNetLemmatizer
# WordNetLemmatizer: Herramienta de NLTK para lematizar palabras (reducirlas a su forma base).

from keras.models import load_model
# load_model: Función de Keras para cargar modelos de aprendizaje profundo previamente entrenados.

import pickle
# pickle: Módulo para serializar y deserializar objetos de Python.

import numpy as np
# numpy: Biblioteca para computación numérica y manejo de arreglos multidimensionales.

import datetime
# datetime: Módulo para trabajar con fechas y horas.

import os
# os: Módulo para interactuar con el sistema operativo (manejo de archivos y directorios).

from openpyxl import Workbook, load_workbook 
# Workbook: Clase de openpyxl para crear y manipular libros de Excel.
# load_workbook: Función de openpyxl para cargar libros de Excel existentes.

import time
# time: Módulo para manejar funciones relacionadas con el tiempo.

from modules import timezones
# timezones: Módulo personalizado para manejar zonas horarias (se asume).

from modules import math
# math: Módulo personalizado para realizar operaciones matemáticas (se asume).

from modules.new_user import registrar_usuario
# registrar_usuario: Función personalizada para registrar un nuevo usuario (se asume).

from modules.captcha import validar_recaptcha
# validar_recaptcha: Función personalizada para validar un CAPTCHA (se asume).

from modules.email_handler_gmail import enviar_correo_gmail
# enviar_correo_gmail: Función personalizada para enviar correos electrónicos usando Gmail (se asume).


app = Flask(__name__)
app.secret_key = '6LcQOQwqAAAAAIiEDhZ0jR_PG8HojjNA0Z0r4s3B'  # Necesario para usar sesiones

# Configuración de Gmail
gmail_user = 'correo@gmail.com' # Reemplazar por correo para chatbot
gmail_password = '' # Incluir contraseña de la cuenta de correo electrónico para chatbot
corporate_email = 'correo@ejemplo.com' # Reemplazar por correo corporativo
logo_path = 'C:\chatbot-captcha\static\logo-final.png'  # Ruta al logo de la empresa

# Cargar modelos y datos necesarios
lemmatizer = WordNetLemmatizer()
words = pickle.load(open('words.pkl', 'rb'))
classes = pickle.load(open('classes.pkl', 'rb'))
model = load_model('chatbot_model.h5')

# Cargar intenciones desde el archivo Excel
def cargar_intenciones_desde_excel():
    archivo_excel = 'preguntas_respuestas.xlsx'
    intenciones = []
    if os.path.exists(archivo_excel):
        workbook = load_workbook(archivo_excel)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            etiqueta, patron, respuesta = row
            if etiqueta and patron and respuesta:
                intenciones.append({
                    "etiqueta": etiqueta,
                    "patrones": [patron],
                    "respuestas": [respuesta]
                })
    return intenciones

intents = {"intents": cargar_intenciones_desde_excel()}

registro_path = 'archivo_traza.xlsx'

if os.path.exists(registro_path):
    registro_libro = load_workbook(registro_path)
    registro_hoja = registro_libro.active
else:
    registro_libro = Workbook()
    registro_hoja = registro_libro.active
    registro_hoja.append(["Fecha y Hora", "Pregunta del Usuario", "Respuesta del Chatbot"])

saludos = ["hola", "buenos días", "buenas tardes", "buenas noches"]

def clean_up_sentence(sentence):
    sentence = sentence.lower()
    sentence_words = nltk.word_tokenize(sentence)
    sentence_words = [lemmatizer.lemmatize(word) for word in sentence_words]
    return sentence_words

def bag_of_words(sentence):
    sentence_words = clean_up_sentence(sentence)
    bag = [0] * len(words)
    for w in sentence_words:
        for i, word in enumerate(words):
            if word == w:
                bag[i] = 1
    return np.array(bag)

def predict_class(sentence):
    bow = bag_of_words(sentence)
    res = model.predict(np.array([bow]))[0]
    ERROR_THRESHOLD = 0.25
    results = [[i, r] for i, r in enumerate(res) if r > ERROR_THRESHOLD]
    results.sort(key=lambda x: x[1], reverse=True)
    return_list = [{'intent': classes[r[0]], 'probability': str(r[1])} for r in results]
    return return_list

def get_response(intents_list, intents_json):
    if intents_list:
        tag = intents_list[0]['intent']
        for intent in intents_json['intents']:
            if intent['etiqueta'] == tag:
                return random.choice(intent['respuestas'])
    return "No se encontró una respuesta adecuada."

def es_solo_saludo(pregunta):
    return pregunta.lower() in saludos

def limpiar_saludo(pregunta):
    palabras = pregunta.lower().split()
    palabras_sin_saludo = [palabra for palabra in palabras if palabra not in saludos]
    return " ".join(palabras_sin_saludo)

MAX_WAIT_TIME = 5  

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_response', methods=['POST'])
def generate_response():
    data = request.get_json()
    pregunta = data['pregunta']
    if pregunta:
        print(f"Pregunta recibida: {pregunta}")

        respuesta = ""
        if any(op in pregunta.lower() for op in ['+', '-', '*', '/']):
            respuesta = math.evaluar_operacion(pregunta, pregunta)
        elif "hora" in pregunta.lower() and "en" in pregunta.lower():
            partes = pregunta.lower().split(" en ")
            if len(partes) > 1:
                pais = partes[1].strip()
                respuesta = timezones.obtener_hora(pais)
            else:
                respuesta = "No se pudo determinar el país para obtener la hora."
        elif "hora" in pregunta.lower():
            respuesta = timezones.obtener_hora_local()
        else:
            if es_solo_saludo(pregunta):
                respuesta = random.choice(["Hola", "Buenos días", "Buenas tardes", "Buenas noches"])
            else:
                pregunta_limpia = limpiar_saludo(pregunta)
                start_time = time.time()
                ints = predict_class(pregunta_limpia)
                respuesta = get_response(ints, intents)
                elapsed_time = time.time() - start_time
                if elapsed_time > MAX_WAIT_TIME:
                    return jsonify({'message': 'Respuesta tardía.'}), 500
                
        now = datetime.datetime.now()
        registro_hoja.append([now.strftime("%Y-%m-%d %H:%M:%S"), pregunta, respuesta])
        registro_libro.save(registro_path)
        
        return jsonify({'respuesta': respuesta}), 200
    else:
        return jsonify({'message': 'Escribe una pregunta primero.'}), 400

@app.route('/register_user', methods=['POST'])
def register_user():
    data = request.get_json()
    nombre = data['nombre']
    email = data['email']
    acepta_politicas = data['acepta_politicas']
    captcha_response = data['captcha_response']
    
    if validar_recaptcha(captcha_response):
        resultado = registrar_usuario(nombre, email, acepta_politicas)
        return jsonify({'resultado': resultado})
    else:
        return jsonify({'resultado': 'Captcha incorrecto. Intente de nuevo.'}), 400

@app.route('/send_email', methods=['POST'])
def send_email():
    data = request.get_json()
    user_email = data['email']
    conversation = data['conversation']
    corporate_email = 'serviciohdtecco@gmail.com'  # Define el correo corporativo

    # Intentar enviar el correo con Gmail
    response_message = enviar_correo_gmail(user_email, conversation, gmail_user, gmail_password, corporate_email, logo_path)

    return jsonify({'message': response_message})

if __name__ == '__main__':
    app.run(host='192.168.1.9', port=5000, debug=True)
