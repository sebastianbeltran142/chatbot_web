import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def registrar_usuario(nombre, email, acepta_politicas):
    archivo_excel = 'datos_usuarios.xlsx'

    if os.path.exists(archivo_excel):
        workbook = load_workbook(archivo_excel)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nombre", "Correo Electrónico", "Fecha y Hora", "Acepta Políticas"])

    # Verificar si el usuario ya está registrado
    for row in sheet.iter_rows(values_only=True):
        if row[1] == email:
            return "Usuario ya registrado. Acceso permitido."

    # Registrar nuevo usuario
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([nombre, email, fecha_hora, acepta_politicas])
    workbook.save(archivo_excel)
    return "Usuario registrado con éxito. Acceso permitido."

# Prueba del módulo
if __name__ == "__main__":
    print(registrar_usuario("Juan Pérez", "juan.perez@example.com", True))
